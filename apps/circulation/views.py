from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test, permission_required
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Count
from datetime import timedelta, datetime
from .models import Loan, Reservation, Fine, LoanRequest, Attendance
from apps.catalog.models import Book, BookCopy
from apps.accounts.models import LibraryUser, StudyRoom, StudyRoomBooking


def is_staff_user(user):
    return user.groups.filter(name='Staff').exists() or user.is_staff


@login_required
def circulation_home(request):
    if is_staff_user(request.user):
        return redirect('circulation:staff_dashboard')
    else:
        return redirect('circulation:patron_dashboard')


@login_required
@user_passes_test(is_staff_user)
def staff_dashboard(request):
    pending_loans = Loan.objects.filter(status='active').select_related('user', 'book_copy__book')
    pending_reservations = Reservation.objects.filter(status='active').select_related('user', 'book')
    pending_borrow_requests = LoanRequest.objects.filter(status='pending').select_related('user', 'book_copy__book')
    overdue_loans = Loan.objects.filter(status='active', due_date__lt=timezone.now()).select_related('user', 'book_copy__book')
    recent_returns = Loan.objects.filter(status='returned').order_by('-return_date')[:10]

    # Attendance stats
    today = timezone.now().date()
    active_attendances = Attendance.objects.filter(status='active')
    today_attendances = Attendance.objects.filter(check_in__date=today)

    context = {
        'pending_loans': pending_loans,
        'pending_reservations': pending_reservations,
        'pending_borrow_requests': pending_borrow_requests,
        'overdue_loans': overdue_loans,
        'recent_returns': recent_returns,
        'active_attendances': active_attendances,
        'today_attendances': today_attendances,
    }
    return render(request, 'circulation/staff_dashboard.html', context)


@login_required
def patron_dashboard(request):
    user = request.user
    current_loans = Loan.objects.filter(user=user, status='active').select_related('book_copy__book')
    active_reservations = Reservation.objects.filter(user=user, status='active').select_related('book')
    pending_borrow_requests = LoanRequest.objects.filter(user=user, status='pending').select_related('book_copy__book')
    loan_history = Loan.objects.filter(user=user).exclude(status='active').order_by('-return_date')[:10]
    unpaid_fines = Fine.objects.filter(loan__user=user, status='unpaid')

    context = {
        'current_loans': current_loans,
        'active_reservations': active_reservations,
        'pending_borrow_requests': pending_borrow_requests,
        'loan_history': loan_history,
        'unpaid_fines': unpaid_fines,
        'total_fines': sum(fine.amount for fine in unpaid_fines),
    }
    return render(request, 'circulation/patron_dashboard.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def checkout_book(request):
    if request.method == 'POST':
        book_copy_id = request.POST.get('book_copy_id')
        user_id = request.POST.get('user_id')

        try:
            book_copy = BookCopy.objects.get(id=book_copy_id, status='available')
            user = get_object_or_404(LibraryUser, id=user_id)

            # Check if user has unpaid fines
            if Fine.objects.filter(loan__user=user, status='unpaid').exists():
                messages.error(request, 'User has unpaid fines. Cannot checkout book.')
                return redirect('circulation:checkout')

            # Create loan
            loan = Loan.objects.create(
                user=user,
                book_copy=book_copy,
                due_date=timezone.now() + timedelta(days=14)  # Default 14 days, will be recalculated
            )
            loan.due_date = loan.calculate_due_date()
            loan.save()

            # Update book copy status
            book_copy.status = 'checked_out'
            book_copy.save()

            messages.success(request, f'Book "{book_copy.book.title}" checked out to {user.get_full_name()}.')
            return redirect('circulation:staff_dashboard')

        except BookCopy.DoesNotExist:
            messages.error(request, 'Book copy not available.')
        except Exception as e:
            messages.error(request, f'Error during checkout: {str(e)}')

    available_copies = BookCopy.objects.filter(status='available').select_related('book')[:50]
    context = {
        'available_copies': available_copies,
    }
    return render(request, 'circulation/checkout.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def return_book(request):
    if request.method == 'POST':
        loan_id = request.POST.get('loan_id')

        try:
            loan = Loan.objects.get(id=loan_id, status='active')
            loan.return_date = timezone.now()
            loan.status = 'returned'
            loan.save()

            # Update book copy status
            loan.book_copy.status = 'available'
            loan.book_copy.save()

            # Check for overdue and create fine if needed
            if loan.due_date < timezone.now():
                days_overdue = (timezone.now().date() - loan.due_date.date()).days
                fine_amount = days_overdue * 1.00  # $1 per day
                Fine.objects.create(
                    loan=loan,
                    amount=fine_amount,
                    reason=f'Overdue return: {days_overdue} days'
                )

            messages.success(request, f'Book "{loan.book_copy.book.title}" returned successfully.')
            return redirect('circulation:staff_dashboard')

        except Loan.DoesNotExist:
            messages.error(request, 'Loan not found.')
        except Exception as e:
            messages.error(request, f'Error during return: {str(e)}')

    active_loans = Loan.objects.filter(status='active').select_related('user', 'book_copy__book')[:50]
    context = {
        'active_loans': active_loans,
    }
    return render(request, 'circulation/return.html', context)


@login_required
@permission_required('circulation.add_reservation')
def reserve_book(request, book_id):
    book = get_object_or_404(Book, id=book_id)
    user = request.user

    # Check if user already has an active reservation for this book
    if Reservation.objects.filter(user=user, book=book, status='active').exists():
        messages.warning(request, 'You already have an active reservation for this book.')
        return redirect('catalog:book_detail', pk=book_id)

    # Check if book is available
    if book.is_available():
        messages.info(request, 'Book is currently available. You can check it out directly.')
        return redirect('catalog:book_detail', pk=book_id)

    # Create reservation
    reservation = Reservation.objects.create(
        user=user,
        book=book,
        expiry_date=timezone.now() + timedelta(days=7)  # Reservation expires in 7 days
    )

    messages.success(request, f'Reservation created for "{book.title}".')
    return redirect('circulation:patron_dashboard')


@login_required
@permission_required('circulation.change_loan')
def renew_loan(request, loan_id):
    loan = get_object_or_404(Loan, id=loan_id, user=request.user, status='active')

    # Check if renewal is allowed (not overdue, within renewal limit)
    if loan.due_date < timezone.now():
        messages.error(request, 'Cannot renew overdue loan.')
        return redirect('circulation:patron_dashboard')

    # Simple renewal: add 14 days
    loan.due_date = loan.due_date + timedelta(days=14)
    loan.save()

    messages.success(request, f'Loan renewed. New due date: {loan.due_date.date()}.')
    return redirect('circulation:patron_dashboard')


@login_required
@permission_required('circulation.view_fine')
def pay_fine(request, fine_id):
    fine = get_object_or_404(Fine, id=fine_id, loan__user=request.user, status='unpaid')

    # For now, just mark as paid (in real app, integrate with payment gateway)
    fine.status = 'paid'
    fine.paid_date = timezone.now()
    fine.save()

    messages.success(request, f'Fine of ${fine.amount} paid successfully.')
    return redirect('circulation:patron_dashboard')


@login_required
def borrow_book(request, book_id):
    """Allow patrons to request to borrow a book."""
    book = get_object_or_404(Book, id=book_id)
    user = request.user

    # Check if user has unpaid fines
    if Fine.objects.filter(loan__user=user, status='unpaid').exists():
        messages.error(request, 'You have unpaid fines. Please pay your fines before borrowing books.')
        return redirect('circulation:patron_dashboard')

    # Check if user already has an active loan for this book
    if Loan.objects.filter(user=user, book_copy__book=book, status='active').exists():
        messages.warning(request, 'You already have this book on loan.')
        return redirect('catalog:book_detail', pk=book_id)

    # Check if user already has a pending request for this book
    if LoanRequest.objects.filter(user=user, book_copy__book=book, status='pending').exists():
        messages.warning(request, 'You already have a pending borrow request for this book.')
        return redirect('catalog:book_detail', pk=book_id)

    # Find available copies
    available_copies = book.copies.filter(status='available')

    if not available_copies.exists():
        messages.error(request, 'No copies of this book are currently available.')
        return redirect('catalog:book_detail', pk=book_id)

    # For now, just pick the first available copy
    # In a more sophisticated system, we could let users choose specific copies
    book_copy = available_copies.first()

    # Create loan request
    loan_request = LoanRequest.objects.create(
        user=user,
        book_copy=book_copy
    )

    messages.success(request, f'Borrow request submitted for "{book.title}". Staff will review your request shortly.')
    return redirect('circulation:patron_dashboard')


@login_required
def cancel_borrow_request(request, request_id):
    """Allow patrons to cancel their pending borrow requests."""
    loan_request = get_object_or_404(LoanRequest, id=request_id, user=request.user, status='pending')

    loan_request.cancel()
    messages.success(request, 'Borrow request cancelled successfully.')
    return redirect('circulation:patron_dashboard')


@login_required
@user_passes_test(lambda u: u.is_superuser)
def approve_borrow_request(request, request_id):
    """Allow admin to approve borrow requests."""
    loan_request = get_object_or_404(LoanRequest, id=request_id, status='pending')

    try:
        loan = loan_request.approve(request.user)
        messages.success(request, f'Borrow request approved. Book "{loan.book_copy.book.title}" is now on loan to {loan.user.get_full_name()}.')
    except ValueError as e:
        messages.error(request, f'Could not approve request: {e}')

    return redirect('circulation:staff_dashboard')


@login_required
@user_passes_test(lambda u: u.is_superuser)
def reject_borrow_request(request, request_id):
    """Allow admin to reject borrow requests."""
    loan_request = get_object_or_404(LoanRequest, id=request_id, status='pending')

    if request.method == 'POST':
        reason = request.POST.get('reason', 'Request rejected by admin')
        loan_request.reject(reason)
        messages.success(request, 'Borrow request rejected.')
        return redirect('circulation:staff_dashboard')

    context = {
        'loan_request': loan_request,
    }
    return render(request, 'circulation/reject_request.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def analytics_dashboard(request):
    """Basic analytics dashboard with key metrics."""
    # Time range (default to last 30 days)
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)

    # Basic metrics
    total_books = Book.objects.active().count()
    total_users = LibraryUser.objects.count()
    active_loans = Loan.objects.filter(status='active').count()
    overdue_loans = Loan.objects.filter(status='active', due_date__lt=timezone.now()).count()
    total_fines = Fine.objects.filter(status='unpaid').count()

    # Recent activity (last 30 days)
    recent_loans = Loan.objects.filter(created_at__gte=start_date).count()
    recent_returns = Loan.objects.filter(return_date__gte=start_date).count()
    recent_registrations = LibraryUser.objects.filter(date_joined__gte=start_date).count()

    # Popular books (by loans in period)
    popular_books = Book.objects.filter(
        copies__loans__created_at__gte=start_date
    ).annotate(
        loan_count=Count('copies__loans')
    ).order_by('-loan_count')[:10]

    # Most active users
    active_users = LibraryUser.objects.filter(
        loans__created_at__gte=start_date
    ).annotate(
        loan_count=Count('loans')
    ).order_by('-loan_count')[:10]

    context = {
        'total_books': total_books,
        'total_users': total_users,
        'active_loans': active_loans,
        'overdue_loans': overdue_loans,
        'total_fines': total_fines,
        'recent_loans': recent_loans,
        'recent_returns': recent_returns,
        'recent_registrations': recent_registrations,
        'popular_books': popular_books,
        'active_users': active_users,
        'days': days,
    }
    return render(request, 'circulation/analytics.html', context)


@login_required
@user_passes_test(is_staff_user)
def register_attendance(request):
    """Register attendance for a visitor."""
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        registration_number = request.POST.get('registration_number')
        full_name = request.POST.get('full_name')
        department = request.POST.get('department')
        faculty = request.POST.get('faculty')
        phone = request.POST.get('phone')
        purpose = request.POST.get('purpose')

        user = None
        if user_id:
            try:
                user = LibraryUser.objects.get(id=user_id)
            except LibraryUser.DoesNotExist:
                pass

        attendance = Attendance.objects.create(
            user=user,
            registration_number=registration_number,
            full_name=full_name,
            department=department,
            faculty=faculty,
            phone=phone,
            purpose=purpose,
        )

        messages.success(request, f'Attendance registered for {full_name}.')
        return redirect('circulation:staff_dashboard')

    # Get recent users for quick selection
    recent_users = LibraryUser.objects.filter(
        membership_type__in=['student', 'faculty', 'staff']
    ).order_by('-last_login')[:20]

    context = {
        'recent_users': recent_users,
    }
    return render(request, 'circulation/register_attendance.html', context)


@login_required
@user_passes_test(is_staff_user)
def checkout_attendance(request, attendance_id):
    """Mark attendance as checked out."""
    attendance = get_object_or_404(Attendance, id=attendance_id, status='active')
    attendance.check_out_visitor()
    messages.success(request, f'{attendance.full_name} checked out successfully.')
    return redirect('circulation:attendance_list')


@login_required
@user_passes_test(is_staff_user)
def attendance_list(request):
    """List attendance records with filtering."""
    attendances = Attendance.objects.all()

    # Filter by date range
    date_filter = request.GET.get('date_filter', 'today')
    if date_filter == 'today':
        start_date = timezone.now().date()
        attendances = attendances.filter(check_in__date=start_date)
    elif date_filter == 'week':
        start_date = timezone.now() - timedelta(days=7)
        attendances = attendances.filter(check_in__gte=start_date)
    elif date_filter == 'month':
        start_date = timezone.now() - timedelta(days=30)
        attendances = attendances.filter(check_in__gte=start_date)
    elif date_filter == 'custom':
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            attendances = attendances.filter(check_in__date__gte=start_date)
        if end_date:
            attendances = attendances.filter(check_in__date__lte=end_date)

    # Filter by status
    status = request.GET.get('status')
    if status:
        attendances = attendances.filter(status=status)

    context = {
        'attendances': attendances,
        'date_filter': date_filter,
        'status': status,
    }
    return render(request, 'circulation/attendance_list.html', context)


@login_required
@user_passes_test(is_staff_user)
def export_attendance_excel(request):
    """Export attendance records to Excel."""
    from openpyxl import Workbook
    from django.http import HttpResponse
    from django.utils import timezone

    # Get filtered attendances
    attendances = Attendance.objects.all()

    date_filter = request.GET.get('date_filter', 'today')
    if date_filter == 'today':
        start_date = timezone.now().date()
        attendances = attendances.filter(check_in__date=start_date)
    elif date_filter == 'week':
        start_date = timezone.now() - timedelta(days=7)
        attendances = attendances.filter(check_in__gte=start_date)
    elif date_filter == 'month':
        start_date = timezone.now() - timedelta(days=30)
        attendances = attendances.filter(check_in__gte=start_date)
    elif date_filter == 'custom':
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date:
            attendances = attendances.filter(check_in__date__gte=start_date)
        if end_date:
            attendances = attendances.filter(check_in__date__lte=end_date)

    status = request.GET.get('status')
    if status:
        attendances = attendances.filter(status=status)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Headers
    headers = [
        'Registration Number', 'Full Name', 'Department', 'Faculty', 'Phone',
        'Purpose', 'Check In', 'Check Out', 'Status'
    ]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Data
    for row_num, attendance in enumerate(attendances, 2):
        ws.cell(row=row_num, column=1, value=attendance.registration_number)
        ws.cell(row=row_num, column=2, value=attendance.full_name)
        ws.cell(row=row_num, column=3, value=attendance.department)
        ws.cell(row=row_num, column=4, value=attendance.faculty)
        ws.cell(row=row_num, column=5, value=attendance.phone)
        ws.cell(row=row_num, column=6, value=attendance.purpose)
        ws.cell(row=row_num, column=7, value=attendance.check_in.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row_num, column=8, value=attendance.check_out.strftime('%Y-%m-%d %H:%M:%S') if attendance.check_out else '')
        ws.cell(row=row_num, column=9, value=attendance.get_status_display())

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"attendance_{date_filter}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'

    wb.save(response)
    return response


@login_required
def room_booking_view(request):
    """Study room booking view."""
    if request.method == 'POST':
        room_id = request.POST.get('room')
        date_str = request.POST.get('date')
        start_time_str = request.POST.get('start_time')
        end_time_str = request.POST.get('end_time')
        participants = request.POST.get('participants')
        purpose = request.POST.get('purpose')
        additional_info = request.POST.get('additional_info')

        try:
            room = StudyRoom.objects.get(id=room_id, is_active=True)
            booking_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            start_time = datetime.strptime(start_time_str, '%H:%M').time()
            end_time = datetime.strptime(end_time_str, '%H:%M').time()
            participants_count = int(participants)

            # Validation
            today = timezone.now().date()
            if booking_date < today:
                messages.error(request, 'Cannot book rooms for past dates.')
                return redirect('circulation:room_booking')

            if participants_count > room.capacity:
                messages.error(request, f'This room can only accommodate {room.capacity} people.')
                return redirect('circulation:room_booking')

            # Check for time conflicts
            conflicting_bookings = StudyRoomBooking.objects.filter(
                room=room,
                date=booking_date,
                status__in=['pending', 'confirmed']
            ).filter(
                Q(start_time__lt=end_time, end_time__gt=start_time)
            )

            if conflicting_bookings.exists():
                messages.error(request, 'This room is not available for the selected time slot.')
                return redirect('circulation:room_booking')

            # Create booking
            booking = StudyRoomBooking.objects.create(
                user=request.user,
                room=room,
                date=booking_date,
                start_time=start_time,
                end_time=end_time,
                number_of_people=participants_count,
                purpose=purpose,
                additional_info=additional_info,
                status='pending'
            )

            messages.success(request, f'Your booking request for {room.name} has been submitted and is pending approval.')
            return redirect('circulation:room_booking')

        except StudyRoom.DoesNotExist:
            messages.error(request, 'Selected room is not available.')
        except ValueError as e:
            messages.error(request, f'Invalid input: {e}')
        except Exception as e:
            messages.error(request, f'An error occurred while processing your booking: {e}')

        return redirect('circulation:room_booking')

    # GET request - show available rooms
    rooms = StudyRoom.objects.filter(is_active=True).order_by('name')
    
    # Calculate occupancy rate and availability
    total_rooms = rooms.count()
    today = timezone.now().date()
    
    # Count today's bookings
    today_bookings = StudyRoomBooking.objects.filter(
        date=today,
        status__in=['pending', 'confirmed']
    ).count()
    
    occupancy_rate = (today_bookings / total_rooms * 100) if total_rooms > 0 else 0

    context = {
        'rooms': rooms,
        'total_rooms': total_rooms,
        'available_rooms': rooms.exclude(
            id__in=StudyRoomBooking.objects.filter(
                date=today,
                status__in=['pending', 'confirmed']
            ).values_list('room_id', flat=True)
        ),
        'occupancy_rate': occupancy_rate,
    }
    return render(request, 'accounts/study_room_booking.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def room_calendar_view(request):
    """Study room booking calendar view for admin."""
    from django.db.models import Q
    
    # Get all bookings for the current month
    today = timezone.now().date()
    first_day_of_month = today.replace(day=1)
    next_month = first_day_of_month.replace(month=first_day_of_month.month + 1) if first_day_of_month.month < 12 else first_day_of_month.replace(year=first_day_of_month.year + 1, month=1)
    
    bookings = StudyRoomBooking.objects.filter(
        date__gte=first_day_of_month,
        date__lt=next_month
    ).select_related('user', 'room').order_by('date', 'start_time')
    
    rooms = StudyRoom.objects.filter(is_active=True).order_by('name')
    
    # Create calendar data
    calendar_data = {}
    current_date = first_day_of_month
    while current_date < next_month:
        calendar_data[current_date] = {}
        for room in rooms:
            calendar_data[current_date][room] = []
        
        current_date += timezone.timedelta(days=1)
    
    # Populate calendar with bookings
    for booking in bookings:
        if booking.date in calendar_data and booking.room in calendar_data[booking.date]:
            calendar_data[booking.date][booking.room].append(booking)
    
    context = {
        'calendar_data': calendar_data,
        'rooms': rooms,
        'current_month': first_day_of_month.strftime('%B %Y'),
        'first_day_of_month': first_day_of_month,
    }
    return render(request, 'circulation/room_calendar.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def approve_room_booking(request, booking_id):
    """Approve a study room booking."""
    booking = get_object_or_404(StudyRoomBooking, id=booking_id, status='pending')
    
    booking.status = 'confirmed'
    booking.confirmed_by = request.user
    booking.confirmed_at = timezone.now()
    booking.save()
    
    # Send confirmation email
    try:
        from django.core.mail import send_mail
        from django.conf import settings
        
        send_mail(
            subject=f"Study Room Booking Confirmed: {booking.room.name}",
            message=f"Dear {booking.user.get_full_name()},\n\n"
                   f"Your study room booking has been confirmed.\n\n"
                   f"Room: {booking.room.name}\n"
                   f"Date: {booking.date.strftime('%B %d, %Y')}\n"
                   f"Time: {booking.start_time.strftime('%I:%M %p')} - {booking.end_time.strftime('%I:%M %p')}\n"
                   f"Purpose: {booking.purpose}\n\n"
                   f"Please arrive on time and present your university ID.\n\n"
                   f"Best regards,\nRamat Library Administration",
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[booking.user.email],
            fail_silently=False,
        )
    except Exception as e:
        # Log error but don't fail the approval
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Failed to send booking confirmation email to {booking.user.username}: {e}")
    
    messages.success(request, f'Booking for {booking.room.name} has been confirmed.')
    return redirect('circulation:room_calendar')


@login_required
@user_passes_test(lambda u: u.is_superuser)
def reject_room_booking(request, booking_id):
    """Reject a study room booking."""
    booking = get_object_or_404(StudyRoomBooking, id=booking_id, status='pending')
    
    if request.method == 'POST':
        reason = request.POST.get('reason', 'Booking rejected by administrator')
        
        booking.status = 'rejected'
        booking.rejected_by = request.user
        booking.rejected_at = timezone.now()
        booking.rejection_reason = reason
        booking.save()
        
        # Send rejection email
        try:
            from django.core.mail import send_mail
            from django.conf import settings
            
            send_mail(
                subject=f"Study Room Booking Rejected: {booking.room.name}",
                message=f"Dear {booking.user.get_full_name()},\n\n"
                       f"Your study room booking has been rejected.\n\n"
                       f"Room: {booking.room.name}\n"
                       f"Date: {booking.date.strftime('%B %d, %Y')}\n"
                       f"Time: {booking.start_time.strftime('%I:%M %p')} - {booking.end_time.strftime('%I:%M %p')}\n"
                       f"Reason: {reason}\n\n"
                       f"Please contact the library administration if you have questions.\n\n"
                       f"Best regards,\nRamat Library Administration",
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[booking.user.email],
                fail_silently=False,
            )
        except Exception as e:
            # Log error but don't fail the rejection
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Failed to send booking rejection email to {booking.user.username}: {e}")
        
        messages.success(request, f'Booking for {booking.room.name} has been rejected.')
        return redirect('circulation:room_calendar')
    
    context = {
        'booking': booking,
    }
    return render(request, 'circulation/reject_room_booking.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def cancel_room_booking(request, booking_id):
    """Cancel a confirmed study room booking."""
    booking = get_object_or_404(StudyRoomBooking, id=booking_id, status='confirmed')
    
    if request.method == 'POST':
        reason = request.POST.get('reason', 'Booking cancelled by administrator')
        
        booking.status = 'cancelled'
        booking.cancelled_by = request.user
        booking.cancelled_at = timezone.now()
        booking.cancellation_reason = reason
        booking.save()
        
        # Send cancellation email
        try:
            from django.core.mail import send_mail
            from django.conf import settings
            
            send_mail(
                subject=f"Study Room Booking Cancelled: {booking.room.name}",
                message=f"Dear {booking.user.get_full_name()},\n\n"
                       f"Your study room booking has been cancelled.\n\n"
                       f"Room: {booking.room.name}\n"
                       f"Date: {booking.date.strftime('%B %d, %Y')}\n"
                       f"Time: {booking.start_time.strftime('%I:%M %p')} - {booking.end_time.strftime('%I:%M %p')}\n"
                       f"Reason: {reason}\n\n"
                       f"We apologize for any inconvenience. Please contact us to reschedule.\n\n"
                       f"Best regards,\nRamat Library Administration",
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[booking.user.email],
                fail_silently=False,
            )
        except Exception as e:
            # Log error but don't fail the cancellation
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Failed to send booking cancellation email to {booking.user.username}: {e}")
        
        messages.success(request, f'Booking for {booking.room.name} has been cancelled.')
        return redirect('circulation:room_calendar')
    
    context = {
        'booking': booking,
    }
    return render(request, 'circulation/cancel_room_booking.html', context)


@login_required
def my_room_bookings(request):
    """View user's own room bookings."""
    bookings = StudyRoomBooking.objects.filter(user=request.user).order_by('-created_at')
    
    context = {
        'bookings': bookings,
    }
    return render(request, 'circulation/my_room_bookings.html', context)


@login_required
def cancel_my_room_booking(request, booking_id):
    """Allow users to cancel their own pending bookings."""
    booking = get_object_or_404(StudyRoomBooking, id=booking_id, user=request.user, status='pending')
    
    booking.status = 'cancelled'
    booking.cancelled_by = request.user
    booking.cancelled_at = timezone.now()
    booking.cancellation_reason = 'User cancelled'
    booking.save()
    
    messages.success(request, f'Your booking for {booking.room.name} has been cancelled.')
    return redirect('circulation:my_room_bookings')
